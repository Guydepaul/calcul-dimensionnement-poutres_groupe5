# Nous allons avoir ici toutes les fonctions pour implementer nos formules
import openpyxl
from back_data_constantes import *
import math

# -------------------- Début Extraction des données dans le fichier excel ---------------------- #
class GetDataExcelFile():
    def __init__(self) -> None:
        self.excel_file = "project/catalogue.xlsx"
        self.load_file = openpyxl.load_workbook(self.excel_file)
    
    def getDataProfilee(self, profile_type):
        worksheet = self.load_file[f"{profile_type}"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
        return excel_data
            
    def getDataAcier(self, sheet):
        worksheet = self.load_file[f"{sheet}"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
        return excel_data
    
    def getDataCourbeFlambement(self, sheet):
        worksheet = self.load_file[f"{sheet}"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
        return excel_data
    
    def getDataCoefficientC1C2C3(self, sheet):
        worksheet = self.load_file[f"{sheet}"]

        excel_data = []
        for row in worksheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
        return excel_data
            
# -------------------- Fin Extraction des données dans le fichier excel ---------------------- #
            
# ---------------------------- Début implementation des formules ----------------------------- #
def calcul_Pn(G, Q, a):
    Pn = ((1.35*G) + (1.5*Q)) * a
    return round(Pn, 2)

def calcul_E1(): # Valeurs dans la catalogue
    E1 = E * (1 - COEFF_POISSON**2)
    return round(E1, 2)

def calcul_a(n0, E1, Pn, t):
    a = (4 * n0 / 15) * (1 + (72*E1 / n0**4 * Pn)) * t
    return round(a, 2)

def calcul_H(FsurL, E1, t):
    H = 1.5 * (PI**2 / 4) * FsurL * E1 * t
    return H

def calcul_C(b, r, tw): # Valeurs dans la catalogue
    C = (b - (2*r*tw)) / 2
    return C

def calcul_Epsilone(fy): # Valeurs dans la catalogue
    epsilone = math.sqrt(235/fy)
    return epsilone

def calcul_Alpha(d, h, Nsd, tw, fy, tf, r): # Autres valeurs dans la catalogue
    alpha = (1/d) * ((h/2) + ((1/2) * (Nsd/2*tw*fy)) - (tf + r))
    return alpha

def calcul_q(G, Q):
    q = (1.35*G) + (1.5*Q)
    return q

def calcul_Msd(q, B):
    Msd = (q * B**2) / 8
    return Msd

def calcul_Vsd(q, B):
    Vsd = (q * B) / 8
    return Vsd

def calcul_Weff(Ieff, Vi): # (Vi ou Vs) Soit on lu dans la cat ou entrer par l'utilisateur
    Weff = Ieff / Vi
    return Weff

def calcul_Mr(W, fy, classe: int): # Autres valeurs dans la catalogue et constantes
    Mr = W*fy / GAMA_M0 if classe in {1, 2, 3} else W*fy / GAMA_M1
    return Mr

def calcul_Vsd_appuis(Q, l, nbre_appuis: int, G=None):
    if nbre_appuis == 2:
        return 1.5 * Q * l / 2
    elif nbre_appuis == 3 and G is not None:
        return 5 * G * l / 8
    else:
        return None

def calcul_Av(type_profile, A: float, B: float, tf: float, tw: float, r: float, h: float=None):
    if type_profile in ['I', 'H']:
        return (A - 2*B*tf) + ((tw + (2*r))*tf)
    elif type_profile == '[':
        return (A - 2*B*tf) + ((tw + r)*tf)
    else:
        return (h-2*tf) * tw

def calcul_Vplrd(Av, fy):
    Vplrd = Av * fy / math.sqrt(3) * GAMA_M0
    return Vplrd

def calcul_rau(Vsd_appuis, Vplrd):
    rau = ((2*Vsd_appuis / Vplrd) - 1)**2
    return rau

def calcul_Frd(rau, fy):
    frd = (1 - rau) * fy
    return frd

def calcul_Mvrd(Wpl, rau, Av, tw, fy):
    Mvrd = (Wpl - (rau * Av**2 / 4*tw)) * (fy / GAMA_M0)
    return Mvrd

def calcul_flexion_deviee(qv, qh, Gv, Gh, alpha, l, Wpl_y, Wpl_z, fy):
    cos_alpha = math.cos(alpha)
    sin_alpha = math.sin(alpha)
    qy = ((qv*cos_alpha) + (Gv*cos_alpha) + (qh*cos_alpha) + (Gh*sin_alpha))
    qz = ((qv*sin_alpha) + (Gv*sin_alpha) + (qh*cos_alpha) + (Gh*cos_alpha))
    
    Qy_ponderee = (((qv*cos_alpha) + (qh*sin_alpha)) * 1.5) + (((Gv*cos_alpha) + (Gh*sin_alpha)) * 1.35) 
    Qz_ponderee = (((qv*sin_alpha) + (qh*cos_alpha)) * 1.5) + (((Gv*sin_alpha) + (Gh*cos_alpha)) * 1.35) 
    
    My_sd = Qy_ponderee*(l**2)/8
    Mz_sd = Qz_ponderee*(l**2)/8
        
    Mpl_y_rd = Wpl_y*fy/GAMA_M0
    Mpl_z_rd = Wpl_z*fy/GAMA_M0
    
    reponse = {
        'qy': qy,
        'qz': qz,
        'Qy_ponderee': Qy_ponderee,
        'Qz_ponderee': Qz_ponderee,
        'My_sd': My_sd,
        'Mz_sd': Mz_sd,
        'Mpl_y_rd': Mpl_y_rd,
        'Mpl_z_rd': Mpl_z_rd,
    }
    
    return reponse

def calcul_Nplrd(A, fy):
    Nplrd = A * fy / GAMA_M0
    return Nplrd

def calcul_n(Nsd, Nplrd):
    n = Nsd / Nplrd
    return n

def calcul_f(Cas, P, L, Iy=None, Iz=None, G=None):
    if Cas == 'CEXP' and Iy is not None:
        return (5 * P * (L**4)) / 384 * E * Iy
    elif Cas == 'PP' and Iz is not None and G is not None:
        return (2.05 * G * ((L/2)**4)) / 384 * E * Iz
    else:
        return None

def calcul_BetaOmega(classe: int, Wply, Wely):
    if classe in {1, 2}:
        return 1
    elif classe == 3 and Wely is not None or classe == 4 and Wely is not None:
        return Wely / Wply
    else:
        return None

def calcul_PhiLT(AlphaLT, GammaBarLT):
    PhiLT = 0.5*(1 + (AlphaLT*(GammaBarLT - 0.2)) + GammaBarLT**2)
    return PhiLT

def calcul_Gama1(Epsilone):
    Gama1 = 93.9 * Epsilone
    return Gama1

def calcul_GamaLT(L, iy, C1, h, tf):
    denom = C1 * (1 + (1/20) * ((L/iy)/(h/tf))**2)
    GamaLT = (L/iy) / math.sqrt(denom)
    return GamaLT

def calcul_GamaBarLT(GamaLT, Gamma1, BetaOmega):
    GamaBarLT = (GamaLT / Gamma1) * BetaOmega**0.5
    return GamaBarLT

def calcul_Xlt(PhiLT, GamaBarLT):
    Xlt = 1 / (PhiLT + (((PhiLT**2) - (GamaBarLT**2))**0.5))
    return Xlt

def calcul_Mbrd(BetaOmega, Xlt, Wely, fy):
    Mbrd = (Xlt * BetaOmega * Wely * fy) / GAMA_M1
    return Mbrd

def calcul_Mcr(Cas, C1, C2, C3, Iy, K, L, Kw, Iw, G, E, It, Zy, Zj, Zg):
    if Cas == 'Cas1':
        Mcr1 = (C1*(PI**2)*Iy/((K*L)**2))
        Mcr2 = (((((K/Kw)**2)*Iw/Iy) + (((K*L)**2)*G*It / (PI**2)*E*Iy) + ((C2*Zy) - (C3*Zj))**2)**0.5) - ((C2*Zg) - (C3*Zj))

        Mcr = Mcr1 * Mcr2
        return Mcr
    
    elif Cas == 'Cas2':
        Mcr1 = (G*(PI**2)*Iy/((K*L)**2))
        Mcr2 = (((((K/Kw)**2)*Iw/Iy) + (((K*L)**2)*G*It / (PI**2)*E*Iy) + ((C2*Zy)**2))**0.5) - ((C2*Zg))

        Mcr = Mcr1 * Mcr2
        return Mcr
    
    elif Cas == 'Cas3':
        Mcr1 = (G*(PI**2)*Iy/((K*L)**2))
        Mcr2 = (((((K/Kw)**2)*Iw/Iy) + (((K*L)**2)*G*It / (PI**2)*E*Iy))**0.5)

        Mcr = Mcr1 * Mcr2
        return Mcr

    elif Cas == 'Cas4':
        Mcr1 = (C1*(PI**2)*Iy/((K*L)**2))
        Mcr2 = (((Iw/Iy) + (((L**2))*G*It / (PI**2)*E*Iy))**0.5)

        Mcr = Mcr1 * Mcr2
        return Mcr
    
    else:
        return None
        
def calcul_BetaT(Ifc, Ift):
    BetaT = Ifc / (Ifc + Ift)
    return BetaT

def calcul_Zy(Za, Zs):
    return (Za - Zs)

def calcul_Zj(BetaT, Cas, hs, hl=None, h=None):
    if Cas == 'Cas1':
        if BetaT > 0.5:
            Zj = 0.8 * ((2*BetaT) - 1) * (hs/2)
        else:
            Zj = ((2*BetaT) - 1) * (hs/2)
    elif Cas == 'Cas2' and None not in [hl, h]:
        if BetaT > 0.5:
            Zj = 0.8 * ((2*BetaT) - 1) * (1 + (hl/h)) * (hs/2)
        else:
            Zj = ((2*BetaT) - 1) * (1 + (hl/h)) * (hs/2)
    else:
        Zj = 0
        
    return Zj

def calcul_hs(hl, tf):
    return (hl - tf)

def calcul_Ktao(Cas, a, d):
    if Cas == 'cas1':
        return 5.34
    elif Cas == 'cas2':
        return 4 + (5.34/((a/d)**2))
    else:
        return 5.34 + (4/((a/d)**2))

def calcul_GamaBar_w(d, tw, Epsilone, Ktao):
    GamaBar_w = (d/tw) / 37.4*Epsilone*math.sqrt(Ktao)
    return GamaBar_w

def calcul_Tao_ba(GamaBar_w, fyw):
    if GamaBar_w <= 0.8:
        return fyw / math.sqrt(3)
    elif GamaBar_w < 1.2:
        return (1 - (0.625*(GamaBar_w - 0.8))) * (fyw / math.sqrt(3))
    else:
        return (0.9*GamaBar_w) * (fyw / math.sqrt(3))

def calcul_Vbrd(d, tw, Tao_ba):
    Vbrd = (d*tw*Tao_ba) / GAMA_M1
    return Vbrd

def calcul_Nfrd(A, fy):
    return A * fy

def calcul_Mfrd(MfrdPoint, Nsd, Nfrd):
    Mfrd = MfrdPoint * (1 - (Nsd/Nfrd))
    return Mfrd

# ---------------------------- Fin implementation des formules ----------------------------- #

# ------------------------- Début implementation des verifications -------------------------- #
def verification_classe_section_poutre(Cas, CsurTf, DsurTw, Epsilone, Alpha=None):
    classe_semelle = None
    classe_ame = None
    
    if CsurTf <= 10*Epsilone:
        classe_semelle = 1
    elif CsurTf <= 11*Epsilone:
        classe_semelle = 2
    elif CsurTf <= 15*Epsilone:
        classe_semelle = 3
    else:
        classe_semelle = 4
        
    if Cas == 'Cas1':
        if DsurTw <= 33*Epsilone:
            classe_ame = 1
        elif DsurTw <= 38*Epsilone:
            classe_ame = 2
        elif DsurTw <= 42*Epsilone:
            classe_ame = 3
        else:
            classe_ame = 4
            
        return f"Classe {max(classe_semelle, classe_ame)}"
    
    if Cas == 'Cas2':
        if DsurTw <= 72*Epsilone:
            classe_ame = 1
        elif DsurTw <= 83*Epsilone:
            classe_ame = 2
        elif DsurTw <= 124*Epsilone:
            classe_ame = 3
        else:
            classe_ame = 4
            
        return f"Classe {max(classe_semelle, classe_ame)}"
    
    if Cas == 'Cas3':
        if Alpha > 0.5:
            comparator = (396*Epsilone) / ((13*Alpha) - 1)
            if DsurTw <= comparator:
                classe_ame = 1
            else:
                comparator = (456*Epsilone) / ((13*Alpha) - 1)
                if DsurTw <= comparator:
                    classe_ame = 2
                
        if Alpha < 0.5:
            comparator = (36*Epsilone) / Alpha
            if DsurTw <= comparator:
                classe_ame = 1
            else:
                comparator = (41.5*Epsilone) / Alpha
                if DsurTw <= comparator:
                    classe_ame = 2
        
        if None not in [classe_semelle, classe_ame]:
            return f"Classe {max(classe_semelle, classe_ame)}"
        
        else:
            return "Classe non définie"

def verification_resistance_poutre(Cas, Classe, Mplrd, Msd, Melrd, Mo_rd, Vpl_rd, Vsd, Mysd, Mzsd, Wy, Wz, Fy, Mpl_y_rd, Mpl_z_rd, n):
    # sourcery skip: assign-if-exp, merge-duplicate-blocks, remove-redundant-if, switch
    if Cas == 'Cas1': # Verification a l'absence de l'effort tranchant
        if Classe in [1, 2]:
            if Mplrd >= Msd:
                return "Condition verifiée"
            else:
                return "Condition non verifiée"
        elif Classe == 3:
            if Melrd >= Msd:
                return "Condition verifiée"
            else:
                return "Condition non verifiée"
        elif Classe == 4:
            if Mo_rd == Msd:
                return "Condition verifiée"
            else:
                return "Condition non verifiée"
        else:
            return "Condition non verifiée"
            
    elif Cas == 'Cas2': # Verification au cisaillement
        if Vpl_rd >= Vsd:
            return "Condition verifiée"
        else:
            return "Condition non verifiée"
    
    elif Cas == 'Cas3': # Verification en presence de l'effort tranchant et du moment flechissant
        if Vsd <= 0.5*Vpl_rd and Msd <= Mplrd:
            return "Condition verifiée"
        elif Vsd > 0.5*Vpl_rd and Msd <= Mplrd:
            return "Condition verifiée"
        else:
            return "Condition non verifiée"
        
    elif Cas == 'Cas4': # Verification en flexion deviee
        if Classe in [1, 2]:
            comparator1 = ((Mysd/Mpl_y_rd)**2) + (Mzsd/Mpl_z_rd)**(5*n)
            if comparator1 <= 1:
                return "Condition verifiée"
            else:
                return "Condition non verifiée"
        elif Classe == 3:
            comparator1 = (Mysd/Wy) + (Mzsd/Wz)
            comparator2 = Fy/GAMA_M0
            if comparator1 <= comparator2:
                return "Condition verifiée"
            else:
                return "Condition non verifiée"
        else:
            return f"Resultat non dispo pour la classe {Classe}"

def verification_rigidite_poutres(Cas, f, l):
    if Cas == 'CEXP':
        comparator = l/200
        return "Condition verifiée" if f <= comparator else "Condition non verifiée"
    elif Cas == 'PP':
        comparator = l/400
        return "Condition verifiée" if f <= comparator else "Condition non verifiée"
    else:
        return f"Resultat non dispo pour le cas {Cas}"
    
def verification_stabilite_deversement_poutres(GamaLT: float):
    if GamaLT <= 0.4:
        return "Stabilité au deversoire assurée"
    else:
        return "Stabilité au deversoire non assurée"
    
def verification_convenance_poutre(Msd, Mb_rd):
    if Msd <= Mb_rd:
        return "Le poutre convient (Condition vérifiée)"
    else:
        return "Le poutre ne convient pas"
    
def verification_resistance_ames_poutres_voilement_cisaillement(Ktao, Epsilone, d, Tw):
    comparator1 = d/Tw
    comparator2 = 30*Epsilone*math.sqrt(Ktao)
    return 1 if comparator1 > comparator2 else 0

def verification_interaction_Vsd_Msd_Nsd(Vsd, Vba_rd, Msd, Mf_rd, Mpl_rd):
    if Vsd <= 0.5*Vba_rd:
        return "Condition vérifiée"
    comparator = Mf_rd + ((Mpl_rd - Mf_rd)*(1 - ((2*Vsd/Vba_rd) - 1)**2))
    return "Condition vérifiée" if Msd <= comparator else "Condition non vérifiée"

def verification_flexion_deviee(classe, Nsd, A, fy, My_sd, Mz_sd, Mpl_y_rd, Mpl_z_rd, Wx, Wy, Wz, Nplrd):
    if classe in [1, 2]:
        n = Nsd/Nplrd
        beta = 5*n
        cond1 = (((My_sd/Mpl_y_rd)**2) + ((Mz_sd/Mpl_z_rd)**beta))
        return 1 if cond1 <= 1 else 0
    elif classe == 3:
        cond2 = (My_sd/Wy) + (Mz_sd/Wz)
        cond3 = fy/GAMA_M0
        fyd = fy/GAMA_M0
        cond4 = (Nsd/A*fyd) + (My_sd/Wx*fyd) + (My_sd/Wy*fyd)
        return 1 if cond2 <= cond3 or cond4 <= 1 else 0
    else:
        return 0

# ------------------------- Fin implementation des verifications -------------------------- #
